# -*- coding: utf-8 -*-
"""
@Time      : 2020/6/17 18:13
@Author    : William.sv@icloud.com
@File      : mergeFile.py
@ Software : PyCharm
@Desc      : 
"""

import os
import csv
import time
import openpyxl


def csv_merge_convert_xlsx(file_path,file_name):
    work_book = openpyxl.Workbook()
    work_sheet = work_book.create_sheet('Sheet')
    row = 1
    lines_data = []
    csv_files = os.listdir(file_path)
    for file in csv_files:
        # 该地址为Windows上的地址，如需在其他系统执行，请改为相应的地址格式
        with open(file_path + '\\' + file, 'r', newline='',encoding='utf-8') as csv_file:
            lines = csv.reader(csv_file)
            for line in lines:
                lines_data.append(line)
    for line in lines_data:
        lin = 1
        for i in line:
            work_sheet.cell(row=row,column=lin).value = i
            lin += 1
        row += 1
    work_book.save(file_path + file_name + '.xlsx')

if __name__ == '__main__':
    file_name = ''
    file_path = ''
    csv_merge_convert_xlsx(file_path,file_name)
